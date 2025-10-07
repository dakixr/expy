from __future__ import annotations

from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .nodes import ColumnNode, RowNode, SheetNode, SpacerNode, TableNode
from .styles import Style, Theme, bold, combine_styles, normalize_hex, to_argb

__all__ = ["render_sheet"]


@dataclass
class EffectiveStyle:
    font_name: str
    font_size: float
    bold: bool
    italic: bool
    text_color: str
    fill_color: str | None
    horizontal_align: str | None
    vertical_align: str | None
    indent: int | None
    wrap_text: bool
    number_format: str | None
    border: str | None
    border_color: str | None


def _resolve(styles: Sequence[Style], theme: Theme) -> EffectiveStyle:
    base_style = Style(font_name=theme.font, font_size=theme.base_size, text_color=theme.default_text_color)
    merged = combine_styles(styles, base=base_style)

    font_name = merged.font_name or theme.font
    if merged.mono:
        font_name = theme.mono_font or font_name
    font_size = merged.font_size if merged.font_size is not None else theme.base_size
    if merged.font_size_delta is not None:
        font_size += merged.font_size_delta

    bold_flag = merged.bold if merged.bold is not None else False
    italic_flag = merged.italic if merged.italic is not None else False

    text_color = merged.text_color
    if merged.text_color_key:
        text_color = theme.colors.get(merged.text_color_key, text_color)
    if text_color is None:
        text_color = theme.default_text_color
    text_color = normalize_hex(text_color)

    fill_color = merged.fill_color
    if merged.fill_color_key:
        fill_color = theme.colors.get(merged.fill_color_key, fill_color)
    fill_color_value = normalize_hex(fill_color) if fill_color else None

    return EffectiveStyle(
        font_name=font_name,
        font_size=font_size,
        bold=bold_flag,
        italic=italic_flag,
        text_color=text_color,
        fill_color=fill_color_value,
        horizontal_align=merged.horizontal_align,
        vertical_align=merged.vertical_align,
        indent=merged.indent,
        wrap_text=merged.wrap_text if merged.wrap_text is not None else False,
        number_format=merged.number_format,
        border=merged.border,
        border_color=normalize_hex(merged.border_color) if merged.border_color else None,
    )


def _default_row_height(theme: Theme) -> float:
    return max(theme.base_size * 1.4, 14.0)


def _update_dimensions(
    *,
    col_widths: dict[int, float],
    row_heights: dict[int, float],
    column_index: int,
    row_index: int,
    value: Any,
    style: EffectiveStyle,
    theme: Theme,
    prefer_height: float | None = None,
) -> None:
    text = "" if value is None else str(value)
    width_hint = max(len(text), 1.0)
    existing_width = col_widths.get(column_index, 0.0)
    col_widths[column_index] = max(existing_width, width_hint)

    base_height = prefer_height if prefer_height is not None else _default_row_height(theme)
    row_heights[row_index] = max(row_heights.get(row_index, 0.0), base_height)


def _apply_style(cell, effective: EffectiveStyle, border_fallback_color: str) -> None:
    cell.font = Font(
        name=effective.font_name,
        size=effective.font_size,
        bold=effective.bold,
        italic=effective.italic,
        color=to_argb(effective.text_color),
    )

    if effective.fill_color:
        color = to_argb(effective.fill_color)
        cell.fill = PatternFill(fill_type="solid", start_color=color, end_color=color)

    align_kwargs: dict[str, Any] = {}
    if effective.horizontal_align:
        align_kwargs["horizontal"] = effective.horizontal_align
    if effective.vertical_align:
        align_kwargs["vertical"] = effective.vertical_align
    if effective.indent is not None:
        align_kwargs["indent"] = effective.indent
    if effective.wrap_text:
        align_kwargs["wrap_text"] = True
    if align_kwargs:
        align_kwargs.setdefault("vertical", "bottom")
        cell.alignment = Alignment(**align_kwargs)
    elif cell.alignment is None and effective.wrap_text:
        cell.alignment = Alignment(wrap_text=True)

    if effective.number_format:
        cell.number_format = effective.number_format

    if effective.border:
        border_color = effective.border_color or border_fallback_color
        side = Side(style=effective.border, color=to_argb(border_color))
        cell.border = Border(left=side, right=side, top=side, bottom=side)


def _render_row(
    ws,
    node: RowNode,
    row_index: int,
    theme: Theme,
    col_widths: dict[int, float],
    row_heights: dict[int, float],
) -> None:
    for column_offset, cell_node in enumerate(node.cells, start=1):
        styles = (*node.styles, *cell_node.styles)
        effective = _resolve(styles, theme)
        target_cell = ws.cell(row=row_index, column=column_offset, value=cell_node.value)
        _apply_style(target_cell, effective, theme.table.border_color)
        _update_dimensions(
            col_widths=col_widths,
            row_heights=row_heights,
            column_index=column_offset,
            row_index=row_index,
            value=cell_node.value,
            style=effective,
            theme=theme,
        )


def _render_column(
    ws,
    node: ColumnNode,
    start_row: int,
    theme: Theme,
    col_widths: dict[int, float],
    row_heights: dict[int, float],
) -> int:
    row_index = start_row
    for cell_node in node.cells:
        styles = (*node.styles, *cell_node.styles)
        effective = _resolve(styles, theme)
        target_cell = ws.cell(row=row_index, column=1, value=cell_node.value)
        _apply_style(target_cell, effective, theme.table.border_color)
        _update_dimensions(
            col_widths=col_widths,
            row_heights=row_heights,
            column_index=1,
            row_index=row_index,
            value=cell_node.value,
            style=effective,
            theme=theme,
        )
        row_index += 1
    return row_index


def _render_table(
    ws,
    node: TableNode,
    start_row: int,
    theme: Theme,
    col_widths: dict[int, float],
    row_heights: dict[int, float],
) -> int:
    table_style = combine_styles(node.styles)
    banded = table_style.table_banded if table_style.table_banded is not None else theme.table.banded
    bordered = table_style.table_bordered if table_style.table_bordered is not None else True
    compact = table_style.table_compact if table_style.table_compact is not None else False
    border_color = table_style.border_color or theme.table.border_color
    border_style = table_style.border or theme.table.border_style

    table_border_style = Style(border=border_style, border_color=border_color) if bordered else None
    stripe_style = Style(fill_color=theme.table.stripe_color) if banded and theme.table.stripe_color else None
    compact_height = theme.table.compact_row_height if compact else None

    current_row = start_row

    def render(
        row_node: RowNode,
        *,
        extra: Sequence[Style] = (),
        prefer_height: float | None = None,
    ) -> None:
        for column_offset, cell_node in enumerate(row_node.cells, start=1):
            style_chain = (*node.styles, *row_node.styles, *extra, *cell_node.styles)
            if table_border_style:
                style_chain = (*style_chain, table_border_style)
            effective = _resolve(style_chain, theme)
            target_cell = ws.cell(row=current_row, column=column_offset, value=cell_node.value)
            _apply_style(target_cell, effective, border_color)
            _update_dimensions(
                col_widths=col_widths,
                row_heights=row_heights,
                column_index=column_offset,
                row_index=current_row,
                value=cell_node.value,
                style=effective,
                theme=theme,
                prefer_height=prefer_height,
            )

    if node.header:
        header_extras: list[Style] = [bold]
        if theme.table.header_bg:
            header_extras.append(Style(fill_color=theme.table.header_bg))
        if theme.table.header_text_color:
            header_extras.append(Style(text_color=theme.table.header_text_color))
        render(node.header, extra=header_extras, prefer_height=compact_height)
        current_row += 1

    for idx, row_node in enumerate(node.rows):
        extras: list[Style] = []
        if stripe_style and idx % 2 == 1:
            extras.append(stripe_style)
        render(row_node, extra=extras, prefer_height=compact_height)
        current_row += 1

    return current_row


def _apply_dimensions(ws, col_widths: Mapping[int, float], row_heights: Mapping[int, float]) -> None:
    for column_index, width in col_widths.items():
        letter = get_column_letter(column_index)
        ws.column_dimensions[letter].width = max(width, 8.0)
    for row_index, height in row_heights.items():
        ws.row_dimensions[row_index].height = height


def render_sheet(ws, node: SheetNode, theme: Theme) -> None:
    col_widths: dict[int, float] = {}
    row_heights: dict[int, float] = {}
    row_cursor = 1

    for item in node.items:
        if isinstance(item, RowNode):
            _render_row(ws, item, row_cursor, theme, col_widths, row_heights)
            row_cursor += 1
        elif isinstance(item, ColumnNode):
            row_cursor = _render_column(ws, item, row_cursor, theme, col_widths, row_heights)
        elif isinstance(item, TableNode):
            row_cursor = _render_table(ws, item, row_cursor, theme, col_widths, row_heights)
        elif isinstance(item, SpacerNode):
            for _ in range(item.rows):
                height = item.height if item.height is not None else _default_row_height(theme)
                row_heights[row_cursor] = max(row_heights.get(row_cursor, 0.0), height)
                row_cursor += 1
        else:  # pragma: no cover - defensive programming
            msg = f"Unsupported sheet item: {type(item)!r}"
            raise TypeError(msg)

    _apply_dimensions(ws, col_widths, row_heights)
